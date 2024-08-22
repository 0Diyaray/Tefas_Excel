from datetime import datetime
from openpyxl import Workbook,load_workbook,workbook
import requests
import os
from bs4 import BeautifulSoup

class Fon:
    def __init__(self, fon_kod):
        self.fonkod = fon_kod
        yüzdelikler = self.web_fon(fon_kod)
        self.aylik = yüzdelikler[0]
        self.ucaylik = yüzdelikler[1]
        self.altiaylik = yüzdelikler[2]
        self.yillik = yüzdelikler[3]

    def web_fon(self,fon_kod):
        url = "https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod=" + fon_kod

        response = requests.get(url)
        page_content = response.text
        soup = BeautifulSoup(page_content, 'html.parser')

        yep = soup.find(class_="price-indicators").get_text(strip=True)
        he = yep.split(" ")

        son = []
        for i in he:
            if "%" in i:
                a = i.replace("Getirisi", "").replace("Son", "").replace("%", "").replace(",", ".")
                if a.strip() == "":
                    son.append("NA")
                else:
                    try:
                        bb = float(a)
                        cc = round(bb, 2)
                        dd = (str(cc) + "%").replace(".", ",")
                        son.append(dd)
                    except ValueError:
                        son.append("NA")
        return son



print("Bilgiler alınıyor...")
start_pos=2

fonlar=["IHK","TAU","BIO","TI4","ACD"]  # Fonların kodlarını bu listeye gir.

fon_list=[]
for i in fonlar:
    fon_list.append(Fon(i)) if Fon(i).aylik+Fon(i).yillik!="0,0%0,0%" else print(f"{i} kodunda bir fon bulunamadı\nDiğer fonlara devam ediliyor...")

#Fonları excel dosyasına yazdırma

print("Tüm veriler alındı excel dosyasına yazdırılıyor...")
now=datetime.now()
pyfolder = os.path.dirname(os.path.abspath(__file__))

file_path = os.path.join(pyfolder, "fons.xlsx")
try:
    workbook = load_workbook(file_path)
except:
    workbook = Workbook()

b=workbook.create_sheet(title="{}-{}-{}".format(now.day,now.month,now.year))
b["A1"],b["A1"].style="Fon",'Input'
b["B1"],b["B1"].style="Aylık",'Input'
b["C1"],b["C1"].style="3 Aylık",'Input'
b["D1"],b["D1"].style="6 Aylık",'Input'
b["E1"],b["E1"].style="Yıllık",'Input'
b["H4"]=datetime.now()
for fon in fon_list:
    b[f"A{start_pos}"],b[f"A{start_pos}"].style= fon.fonkod,'Good'
    b[f"B{start_pos}"]= float(fon.aylik.replace(",",".").strip("%"))/100 if fon.aylik != "NA" else "NA"
    b[f"C{start_pos}"]= float(fon.ucaylik.replace(",",".").strip("%"))/100 if fon.ucaylik != "NA" else "NA"
    b[f"D{start_pos}"]= float(fon.altiaylik.replace(",",".").strip("%"))/100 if fon.altiaylik != "NA" else "NA"
    b[f"E{start_pos}"]= float(fon.yillik.replace(",",".").strip("%"))/100 if fon.yillik != "NA" else "NA"
    start_pos+=1
b[f"A{start_pos}"],b[f"A{start_pos}"].style = "Ortalama",'Bad'
b[f"B{start_pos}"] = f"==AVERAGE(B1:B{start_pos-1})"
b[f"C{start_pos}"] = f"==AVERAGE(C1:C{start_pos-1})"
b[f"D{start_pos}"] = f"==AVERAGE(D1:D{start_pos-1})"
b[f"E{start_pos}"] = f"==AVERAGE(E1:E{start_pos-1})"

workbook.save(file_path)
print("Fonların performansları excel tablosuna aktarıldı.")
print("Tamamlandı.",datetime.now())

